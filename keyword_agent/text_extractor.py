"""Text extractor: pulls plain text from structured document formats."""

import logging
import re
import unicodedata
from pathlib import Path

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".eml", ".msg", ".txt"}


def extract(source_path: Path, output_dir: Path) -> dict:
    """Extract text from *source_path* and write a ``.txt`` file to *output_dir*.

    Args:
        source_path: Path to the source document.
        output_dir: Directory where the extracted ``.txt`` file is written.

    Returns:
        A dict with ``output_file``, ``text``, and ``error`` keys.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    ext = source_path.suffix.lower()

    extractors = {
        ".pdf": _extract_pdf,
        ".docx": _extract_docx,
        ".xlsx": _extract_xlsx,
        ".eml": _extract_eml,
        ".msg": _extract_msg,
        ".txt": _extract_txt,
    }

    fn = extractors.get(ext)
    if fn is None:
        return {
            "output_file": None,
            "text": "",
            "error": f"Unsupported extension: {ext}",
        }

    try:
        text = fn(source_path)
        text = _normalize(text)
        output_file = output_dir / (source_path.stem + "_extracted.txt")
        output_file.write_text(text, encoding="utf-8")
        logger.info("Extracted %d chars from %s", len(text), source_path.name)
        return {"output_file": output_file, "text": text, "error": None}
    except Exception as exc:  # pylint: disable=broad-except
        logger.error("Extraction failed for %s: %s", source_path, exc)
        return {"output_file": None, "text": "", "error": str(exc)}


# ---------------------------------------------------------------------------
# Format-specific extractors
# ---------------------------------------------------------------------------

def _extract_pdf(path: Path) -> str:
    import pdfplumber  # noqa: PLC0415

    parts: list[str] = []
    with pdfplumber.open(str(path)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            if page_text.strip():
                parts.append(page_text)
    return "\n\n".join(parts)


def _extract_docx(path: Path) -> str:
    from docx import Document  # noqa: PLC0415

    doc = Document(str(path))
    return "\n".join(para.text for para in doc.paragraphs)


def _extract_xlsx(path: Path) -> str:
    from openpyxl import load_workbook  # noqa: PLC0415

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    rows: list[str] = []
    for sheet in wb.worksheets:
        rows.append(f"[Sheet: {sheet.title}]")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            rows.append("\t".join(cells))
    return "\n".join(rows)


def _extract_eml(path: Path) -> str:
    import email  # noqa: PLC0415
    from email import policy  # noqa: PLC0415

    raw = path.read_bytes()
    msg = email.message_from_bytes(raw, policy=policy.default)
    parts: list[str] = []

    subject = msg.get("subject", "")
    sender = msg.get("from", "")
    recipient = msg.get("to", "")
    if subject:
        parts.append(f"Subject: {subject}")
    if sender:
        parts.append(f"From: {sender}")
    if recipient:
        parts.append(f"To: {recipient}")

    for part in msg.walk():
        content_type = part.get_content_type()
        if content_type in ("text/plain", "text/html"):
            try:
                text = part.get_content()
                if text:
                    parts.append(str(text))
            except Exception:  # pylint: disable=broad-except
                pass

    return "\n".join(parts)


def _extract_msg(path: Path) -> str:
    import extract_msg  # noqa: PLC0415

    with extract_msg.Message(str(path)) as msg:
        parts: list[str] = []
        if msg.subject:
            parts.append(f"Subject: {msg.subject}")
        if msg.sender:
            parts.append(f"From: {msg.sender}")
        if msg.to:
            parts.append(f"To: {msg.to}")
        if msg.body:
            parts.append(msg.body)
        return "\n".join(parts)


def _extract_txt(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


# ---------------------------------------------------------------------------
# Normalisation
# ---------------------------------------------------------------------------

def _normalize(text: str) -> str:
    """Normalise unicode and collapse excessive whitespace."""
    text = unicodedata.normalize("NFKC", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()
